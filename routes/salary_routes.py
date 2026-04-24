import io
import os
from calendar import month_name, monthrange
from datetime import date, datetime
from functools import wraps

import openpyxl
from flask import Blueprint, current_app, flash, redirect, render_template, request, send_file, session, url_for, make_response

from database.db_connection import get_db_connection
from module_access import module_required

salary_bp = Blueprint("salary", __name__, url_prefix="/company")


def company_required(function):
    @wraps(function)
    def wrapper(*args, **kwargs):
        if session.get("role") != "company":
            flash("Please log in as company to access salary management.", "error")
            return redirect(url_for("auth.role_login", role="company"))
        return function(*args, **kwargs)

    return wrapper


def _to_amount(value, default=0.0):
    try:
        cleaned = str(value).replace(",", "").strip()
        if cleaned == "":
            return default
        return round(float(cleaned), 2)
    except (TypeError, ValueError):
        return default


def _build_salary_payload(company, employee, salary_row):
    month_label = month_name[int(salary_row["month"])]
    return {
        "company": company,
        "employee": employee,
        "salary": salary_row,
        "month_label": month_label,
        "generated_on": datetime.now(),
    }


def _render_salary_pdf(payload, salary_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    uploads_dir = os.path.join(current_app.root_path, "static", "uploads", "salary_slips")
    os.makedirs(uploads_dir, exist_ok=True)

    filename = f"salary_slip_{salary_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    absolute_path = os.path.join(uploads_dir, filename)
    relative_path = f"uploads/salary_slips/{filename}"

    company = payload["company"]
    employee = payload["employee"]
    salary = payload["salary"]

    pdf = canvas.Canvas(absolute_path, pagesize=A4)
    width, height = A4

    y = height - 20 * mm
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawString(20 * mm, y, str(company.get("company_name", "Company")))

    pdf.setFont("Helvetica", 10)
    y -= 6 * mm
    pdf.drawString(20 * mm, y, f"Address: {company.get('address') or 'N/A'}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Email: {company.get('email') or 'N/A'}")

    y -= 10 * mm
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(20 * mm, y, "Salary Slip")
    pdf.setFont("Helvetica", 10)
    y -= 6 * mm
    pdf.drawString(20 * mm, y, f"Period: {payload['month_label']} {salary['year']}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Generated: {payload['generated_on'].strftime('%d-%m-%Y %I:%M %p')}")

    y -= 10 * mm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(20 * mm, y, "Employee Details")
    pdf.setFont("Helvetica", 10)
    y -= 6 * mm
    pdf.drawString(20 * mm, y, f"Name: {employee['name']}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Employee ID: {employee['emp_id']}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Department: {employee.get('department') or 'N/A'}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Role: {employee.get('role') or 'N/A'}")

    y -= 10 * mm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(20 * mm, y, "Attendance Summary")
    pdf.setFont("Helvetica", 10)
    y -= 6 * mm
    pdf.drawString(20 * mm, y, f"Total days in month: {salary['total_days']}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Present days: {salary['present_days']}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Absent days: {max(int(salary['total_days']) - int(salary['present_days']), 0)}")

    y -= 10 * mm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(20 * mm, y, "Earnings")
    pdf.setFont("Helvetica", 10)
    y -= 6 * mm
    pdf.drawString(20 * mm, y, f"Basic Salary: {salary['basic_salary']:.2f}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"HRA: {salary['hra']:.2f}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Bonus: {salary['bonus']:.2f}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Gross Salary: {salary['gross_salary']:.2f}")

    y -= 10 * mm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(20 * mm, y, "Deductions")
    pdf.setFont("Helvetica", 10)
    y -= 6 * mm
    pdf.drawString(20 * mm, y, f"PF (Employee): {salary['pf_employee']:.2f}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"Other Deductions: {salary['deductions']:.2f}")
    y -= 5 * mm
    pdf.drawString(20 * mm, y, f"PF (Employer Contribution): {salary['pf_employer']:.2f}")

    y -= 10 * mm
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(20 * mm, y, f"Net Pay: {salary['net_salary']:.2f}")

    pdf.showPage()
    pdf.save()

    return absolute_path, relative_path


@salary_bp.route("/salary/generate", methods=["GET"])
@company_required
@module_required("salary_slip_management")
def salary_generate_page():
    company_id = session.get("company_id")
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT id, name, emp_id, department, role, basic_salary
        FROM employees
        WHERE company_id = %s AND status IN ('active', 'registered', 'pending')
        ORDER BY name ASC
        """,
        (company_id,),
    )
    employees = cursor.fetchall()

    cursor.execute(
        "SELECT id, name FROM departments WHERE company_id = %s ORDER BY name ASC",
        (company_id,),
    )
    departments = cursor.fetchall()

    cursor.close()
    connection.close()

    today = date.today()
    return render_template(
        "company/salary_generate.html",
        employees=employees,
        departments=departments,
        current_month=today.month,
        current_year=today.year,
    )


@salary_bp.route("/generate_salary", methods=["POST"])
@company_required
@module_required("salary_slip_management")
def generate_salary():
    company_id = session.get("company_id")

    employee_id = request.form.get("employee_id", type=int)
    month = request.form.get("month", type=int)
    year = request.form.get("year", type=int)
    salary_mode = request.form.get("salary_mode", "fixed").strip().lower()
    bonus = _to_amount(request.form.get("bonus"), 0.0)
    deductions = _to_amount(request.form.get("deductions"), 0.0)
    per_day_rate = _to_amount(request.form.get("per_day_rate"), 0.0)

    if not employee_id or not month or not year:
        flash("Employee, month, and year are required.", "error")
        return redirect(url_for("salary.salary_generate_page"))

    if month < 1 or month > 12:
        flash("Month must be between 1 and 12.", "error")
        return redirect(url_for("salary.salary_generate_page"))

    if year < 2000 or year > 2100:
        flash("Year is out of supported range.", "error")
        return redirect(url_for("salary.salary_generate_page"))

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT id, name, emp_id, department, role, basic_salary FROM employees WHERE id = %s AND company_id = %s",
            (employee_id, company_id),
        )
        employee = cursor.fetchone()
        if not employee:
            flash("Employee not found for this company.", "error")
            return redirect(url_for("salary.salary_generate_page"))

        cursor.execute("SELECT id, company_name, address, email FROM companies WHERE id = %s", (company_id,))
        company = cursor.fetchone() or {}

        salary_row = _compute_and_save_salary(cursor, connection, company_id, company, employee, month, year, salary_mode, bonus, deductions, per_day_rate)
        if not salary_row:
            flash("Unable to generate salary. Please try again.", "error")
            return redirect(url_for("salary.salary_generate_page"))

        connection.commit()
        saved_salary_id = salary_row["id"]

    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()

    flash("Salary generated successfully and salary slip PDF created.", "success")
    return redirect(url_for("salary.salary_slip_view", salary_id=saved_salary_id))


def _compute_and_save_salary(cursor, connection, company_id, company, employee, month, year, salary_mode, bonus, deductions, per_day_rate):
    """Compute salary for one employee and upsert into DB. Returns salary_row or None."""
    total_days = monthrange(year, month)[1]
    configured_basic = _to_amount(employee.get("basic_salary"), 0.0)
    per_day_rate_calc = round(configured_basic / total_days, 2) if total_days > 0 else 0.0

    # Attendance count
    cursor.execute(
        """
        SELECT COUNT(*) AS present_days FROM attendance
        WHERE company_id = %s AND employee_id = %s
          AND YEAR(date) = %s AND MONTH(date) = %s
          AND LOWER(TRIM(status)) IN ('present', 'late')
        """,
        (company_id, employee["emp_id"], year, month),
    )
    present_days = int((cursor.fetchone() or {}).get("present_days") or 0)
    if present_days == 0:
        cursor.execute(
            """
            SELECT COUNT(*) AS present_days FROM attendance
            WHERE company_id = %s AND employee_id = %s
              AND YEAR(date) = %s AND MONTH(date) = %s
              AND LOWER(TRIM(status)) IN ('present', 'late')
            """,
            (company_id, str(employee["id"]), year, month),
        )
        present_days = max(present_days, int((cursor.fetchone() or {}).get("present_days") or 0))

    if salary_mode == "per_day":
        rate = per_day_rate if per_day_rate > 0 else per_day_rate_calc
        basic_salary = round(rate * present_days, 2)
    else:
        basic_salary = configured_basic

    hra = round(basic_salary * 0.20, 2)
    pf_employee = round(basic_salary * 0.12, 2)
    pf_employer = round(basic_salary * 0.12, 2)
    gross_salary = round(basic_salary + hra + bonus, 2)
    net_salary = round(gross_salary - (pf_employee + deductions), 2)

    cursor.execute(
        """
        INSERT INTO salary (
            employee_id, company_id, month, year, total_days, present_days,
            basic_salary, hra, bonus, pf_employee, pf_employer, deductions,
            gross_salary, net_salary
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
            total_days=VALUES(total_days), present_days=VALUES(present_days),
            basic_salary=VALUES(basic_salary), hra=VALUES(hra), bonus=VALUES(bonus),
            pf_employee=VALUES(pf_employee), pf_employer=VALUES(pf_employer),
            deductions=VALUES(deductions), gross_salary=VALUES(gross_salary),
            net_salary=VALUES(net_salary), updated_at=CURRENT_TIMESTAMP
        """,
        (employee["id"], company_id, month, year, total_days, present_days,
         basic_salary, hra, bonus, pf_employee, pf_employer, deductions, gross_salary, net_salary),
    )

    cursor.execute(
        """
        SELECT s.*, e.name, e.emp_id FROM salary s
        INNER JOIN employees e ON e.id = s.employee_id
        WHERE s.employee_id = %s AND s.company_id = %s AND s.month = %s AND s.year = %s
        """,
        (employee["id"], company_id, month, year),
    )
    salary_row = cursor.fetchone()
    if not salary_row:
        return None

    payload = _build_salary_payload(company, employee, salary_row)
    _, relative_pdf_path = _render_salary_pdf(payload, salary_row["id"])

    cursor.execute(
        """
        INSERT INTO salary_slips (salary_id, pdf_path) VALUES (%s, %s)
        ON DUPLICATE KEY UPDATE pdf_path=VALUES(pdf_path), generated_at=CURRENT_TIMESTAMP
        """,
        (salary_row["id"], relative_pdf_path),
    )
    return salary_row


@salary_bp.route("/generate_salary_department", methods=["POST"])
@company_required
@module_required("salary_slip_management")
def generate_salary_department():
    company_id = session.get("company_id")
    department_id = request.form.get("department_id", type=int)
    month = request.form.get("month", type=int)
    year = request.form.get("year", type=int)
    salary_mode = request.form.get("salary_mode", "fixed").strip().lower()
    bonus = _to_amount(request.form.get("bonus"), 0.0)
    deductions = _to_amount(request.form.get("deductions"), 0.0)
    per_day_rate = _to_amount(request.form.get("per_day_rate"), 0.0)

    if not department_id or not month or not year:
        flash("Department, month, and year are required.", "error")
        return redirect(url_for("salary.salary_generate_page"))

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, name FROM departments WHERE id = %s AND company_id = %s", (department_id, company_id))
        dept = cursor.fetchone()
        if not dept:
            flash("Department not found.", "error")
            return redirect(url_for("salary.salary_generate_page"))

        cursor.execute(
            """
            SELECT id, name, emp_id, department, role, basic_salary
            FROM employees
            WHERE company_id = %s AND department = %s AND status IN ('active', 'registered', 'pending')
            ORDER BY name ASC
            """,
            (company_id, dept["name"]),
        )
        employees = cursor.fetchall()

        if not employees:
            flash(f"No active employees found in department '{dept['name']}'.", "error")
            return redirect(url_for("salary.salary_generate_page"))

        cursor.execute("SELECT id, company_name, address, email FROM companies WHERE id = %s", (company_id,))
        company = cursor.fetchone() or {}

        count = 0
        for emp in employees:
            result = _compute_and_save_salary(cursor, connection, company_id, company, emp, month, year, salary_mode, bonus, deductions, per_day_rate)
            if result:
                count += 1

        connection.commit()

    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()

    flash(f"Salary generated for {count} employee(s) in '{dept['name']}' department.", "success")
    return redirect(url_for("salary.salary_records"))


@salary_bp.route("/salary_records", methods=["GET"])
@company_required
@module_required("salary_slip_management")
def salary_records():
    company_id = session.get("company_id")

    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)

    query = """
        SELECT s.id, s.month, s.year, s.total_days, s.present_days,
               s.basic_salary, s.hra, s.bonus, s.pf_employee, s.pf_employer,
               s.deductions, s.gross_salary, s.net_salary, s.created_at,
               e.name AS employee_name, e.emp_id,
               ss.pdf_path, ss.generated_at
        FROM salary s
        INNER JOIN employees e ON e.id = s.employee_id
        LEFT JOIN salary_slips ss ON ss.salary_id = s.id
        WHERE s.company_id = %s
    """
    params = [company_id]

    if month:
        query += " AND s.month = %s"
        params.append(month)
    if year:
        query += " AND s.year = %s"
        params.append(year)

    query += " ORDER BY s.year DESC, s.month DESC, e.name ASC"

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(query, tuple(params))
    records = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "company/salary_records.html",
        records=records,
        selected_month=month,
        selected_year=year,
    )


@salary_bp.route("/salary_slip/<int:salary_id>", methods=["GET"])
@company_required
@module_required("salary_slip_management")
def salary_slip_view(salary_id):
    company_id = session.get("company_id")

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT s.*, e.name AS employee_name, e.emp_id, e.department, e.role,
               c.company_name, c.address, c.email,
               ss.pdf_path, ss.generated_at
        FROM salary s
        INNER JOIN employees e ON e.id = s.employee_id
        INNER JOIN companies c ON c.id = s.company_id
        LEFT JOIN salary_slips ss ON ss.salary_id = s.id
        WHERE s.id = %s AND s.company_id = %s
        """,
        (salary_id, company_id),
    )
    salary_data = cursor.fetchone()

    cursor.close()
    connection.close()

    if not salary_data:
        flash("Salary record not found.", "error")
        return redirect(url_for("salary.salary_records"))

    return render_template("company/salary_slip_view.html", salary=salary_data, month_name=month_name)


@salary_bp.route("/download_salary/<int:salary_id>", methods=["GET"])
@company_required
@module_required("salary_slip_management")
def download_salary(salary_id):
    company_id = session.get("company_id")

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT s.*, e.id AS employee_pk, e.name AS employee_name, e.emp_id, e.department, e.role,
               c.company_name, c.address, c.email,
               ss.pdf_path
        FROM salary s
        INNER JOIN employees e ON e.id = s.employee_id
        INNER JOIN companies c ON c.id = s.company_id
        LEFT JOIN salary_slips ss ON ss.salary_id = s.id
        WHERE s.id = %s AND s.company_id = %s
        """,
        (salary_id, company_id),
    )
    salary_data = cursor.fetchone()

    if not salary_data:
        cursor.close()
        connection.close()
        flash("Salary record not found.", "error")
        return redirect(url_for("salary.salary_records"))

    pdf_path = salary_data.get("pdf_path")
    absolute_path = os.path.join(current_app.root_path, "static", pdf_path) if pdf_path else None

    if not absolute_path or not os.path.exists(absolute_path):
        employee = {
            "name": salary_data["employee_name"],
            "emp_id": salary_data["emp_id"],
            "department": salary_data.get("department"),
            "role": salary_data.get("role"),
        }
        company = {
            "company_name": salary_data["company_name"],
            "address": salary_data.get("address"),
            "email": salary_data.get("email"),
        }
        payload = {
            "company": company,
            "employee": employee,
            "salary": salary_data,
            "month_label": month_name[int(salary_data["month"])],
            "generated_on": datetime.now(),
        }

        absolute_path, relative_path = _render_salary_pdf(payload, salary_id)
        cursor.execute(
            """
            INSERT INTO salary_slips (salary_id, pdf_path)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE
                pdf_path = VALUES(pdf_path),
                generated_at = CURRENT_TIMESTAMP
            """,
            (salary_id, relative_path),
        )
        connection.commit()

    cursor.close()
    connection.close()

    return send_file(absolute_path, as_attachment=True, download_name=f"salary_slip_{salary_id}.pdf")


@salary_bp.route("/download_sample_excel", methods=["GET"])
@company_required
@module_required("salary_slip_management")
def download_sample_excel():
    """Download sample Excel template for bulk salary generation."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Data"

    headers = ["Employee ID", "Name", "Basic Salary", "Present Days", "Absent Days", "Bonus", "Deduction"]
    ws.append(headers)

    ws.append(["EMP001", "John Doe", "30000", "22", "8", "2000", "500"])
    ws.append(["EMP002", "Jane Smith", "35000", "25", "5", "1500", "300"])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers["Content-Disposition"] = "attachment; filename=salary_template.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@salary_bp.route("/generate_salary_excel", methods=["POST"])
@company_required
@module_required("salary_slip_management")
def generate_salary_excel():
    """Generate salary slips from uploaded Excel file."""
    company_id = session.get("company_id")

    if "salary_excel" not in request.files:
        flash("No file uploaded.", "error")
        return redirect(url_for("salary.salary_generate_page"))

    file = request.files["salary_excel"]
    if file.filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("salary.salary_generate_page"))

    if not file.filename.endswith((".xlsx", ".xls")):
        flash("Invalid file format. Please upload an Excel file (.xlsx or .xls).", "error")
        return redirect(url_for("salary.salary_generate_page"))

    month = request.form.get("excel_month", type=int)
    year = request.form.get("excel_year", type=int)

    if not month or not year:
        flash("Month and year are required.", "error")
        return redirect(url_for("salary.salary_generate_page"))

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required_columns = ["Employee ID", "Name", "Basic Salary", "Present Days", "Absent Days", "Bonus", "Deduction"]

        for col in required_columns:
            if col not in headers:
                flash(f"Missing required column: {col}", "error")
                return redirect(url_for("salary.salary_generate_page"))

        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)

        cursor.execute("SELECT id, company_name, address, email FROM companies WHERE id = %s", (company_id,))
        company = cursor.fetchone() or {}

        total_days = monthrange(year, month)[1]
        success_count = 0
        error_list = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            try:
                emp_id = str(row[headers.index("Employee ID")]).strip()
                name = str(row[headers.index("Name")]).strip()
                basic_salary = _to_amount(row[headers.index("Basic Salary")], 0.0)
                present_days = int(row[headers.index("Present Days")] or 0)
                absent_days = int(row[headers.index("Absent Days")] or 0)
                bonus = _to_amount(row[headers.index("Bonus")], 0.0)
                deduction = _to_amount(row[headers.index("Deduction")], 0.0)

                cursor.execute(
                    "SELECT id, name, emp_id, department, role FROM employees WHERE emp_id = %s AND company_id = %s",
                    (emp_id, company_id),
                )
                employee = cursor.fetchone()

                if not employee:
                    error_list.append(f"Row {row_idx}: Employee ID '{emp_id}' not found")
                    continue

                hra = round(basic_salary * 0.20, 2)
                pf_employee = round(basic_salary * 0.12, 2)
                pf_employer = round(basic_salary * 0.12, 2)
                gross_salary = round(basic_salary + hra + bonus, 2)
                net_salary = round(gross_salary - (pf_employee + deduction), 2)

                cursor.execute(
                    """
                    INSERT INTO salary (
                        employee_id, company_id, month, year, total_days, present_days,
                        basic_salary, hra, bonus, pf_employee, pf_employer, deductions,
                        gross_salary, net_salary
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                        total_days=VALUES(total_days), present_days=VALUES(present_days),
                        basic_salary=VALUES(basic_salary), hra=VALUES(hra), bonus=VALUES(bonus),
                        pf_employee=VALUES(pf_employee), pf_employer=VALUES(pf_employer),
                        deductions=VALUES(deductions), gross_salary=VALUES(gross_salary),
                        net_salary=VALUES(net_salary), updated_at=CURRENT_TIMESTAMP
                    """,
                    (employee["id"], company_id, month, year, total_days, present_days,
                     basic_salary, hra, bonus, pf_employee, pf_employer, deduction, gross_salary, net_salary),
                )

                cursor.execute(
                    """
                    SELECT s.*, e.name, e.emp_id FROM salary s
                    INNER JOIN employees e ON e.id = s.employee_id
                    WHERE s.employee_id = %s AND s.company_id = %s AND s.month = %s AND s.year = %s
                    """,
                    (employee["id"], company_id, month, year),
                )
                salary_row = cursor.fetchone()

                if salary_row:
                    payload = _build_salary_payload(company, employee, salary_row)
                    _, relative_pdf_path = _render_salary_pdf(payload, salary_row["id"])

                    cursor.execute(
                        """
                        INSERT INTO salary_slips (salary_id, pdf_path) VALUES (%s, %s)
                        ON DUPLICATE KEY UPDATE pdf_path=VALUES(pdf_path), generated_at=CURRENT_TIMESTAMP
                        """,
                        (salary_row["id"], relative_pdf_path),
                    )
                    success_count += 1

            except Exception as e:
                error_list.append(f"Row {row_idx}: {str(e)}")
                continue

        connection.commit()
        cursor.close()
        connection.close()

        if error_list:
            flash(f"Processed {success_count} records. Errors: {'; '.join(error_list[:5])}", "warning")
        else:
            flash(f"Successfully generated salary slips for {success_count} employees from Excel.", "success")

        return redirect(url_for("salary.salary_records"))

    except Exception as e:
        flash(f"Error processing Excel file: {str(e)}", "error")
        return redirect(url_for("salary.salary_generate_page"))
